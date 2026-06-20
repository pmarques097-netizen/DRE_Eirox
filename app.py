# -*- coding: utf-8 -*-
"""
Eirox DRE Online Premium
VERSÃO DINÂMICA TOTAL - DRE EIROX ENTERPRISE PREMIUM v1.5

Esta versão recalcula o DRE sempre que o app é executado, lendo diretamente as pastas:
- CONTAS A PAGAR - DRE
- PLANO DE CONTAS - LEGENDA
- POSIÇÃO DE ESTOQUE
- VENDA POR PAGAMENTO

O arquivo DRE_Consolidado_Moderno.xlsx é usado apenas como fallback quando as pastas não forem encontradas.

Como executar:
    streamlit run app.py
"""

from __future__ import annotations

import base64
import hashlib
import hmac
import io
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st

# =========================================================
# CONFIGURAÇÃO
# =========================================================
st.set_page_config(
    page_title="Eirox DRE Online",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_DIR = Path(__file__).resolve().parent

POSSIVEIS_BASES = [
    APP_DIR / "data" / "DRE_Consolidado_Moderno.xlsx",
    APP_DIR / "DRE_Consolidado_Moderno.xlsx",
    Path.cwd() / "data" / "DRE_Consolidado_Moderno.xlsx",
    Path.cwd() / "DRE_Consolidado_Moderno.xlsx",
    Path.home() / "Desktop" / "Dre" / "data" / "DRE_Consolidado_Moderno.xlsx",
    Path.home() / "Desktop" / "Dre" / "DRE_Consolidado_Moderno.xlsx",
]

POSSIVEIS_LOGOS = [
    APP_DIR / "assets" / "logo_eirox.png",
    APP_DIR / "assets" / "logo eirox(3).png",
    APP_DIR / "assets" / "logo eirox.png",
    APP_DIR / "logo_eirox.png",
    APP_DIR / "logo eirox(3).png",
    Path.cwd() / "assets" / "logo_eirox.png",
    Path.cwd() / "assets" / "logo eirox(3).png",
    Path.cwd() / "logo_eirox.png",
]

MESES_ORDEM = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

MES_NUM_TO_ABR = {
    1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
    7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez",
}

# =========================================================
# CONTROLE DE ACESSO
# =========================================================
USUARIOS = {
    "paulomarques": {"senha": "031730", "perfil": "admin", "nome": "Paulo Marques"},
    "admin": {"senha": "031730", "perfil": "admin", "nome": "Administrador"},
    "ubiratan": {"senha": "031730", "perfil": "visualizacao", "nome": "Ubiratan"},
    "vanderlei": {"senha": "031730", "perfil": "visualizacao", "nome": "Vanderlei"},
}

# Mantém o login ativo no mesmo navegador mesmo após recarregar/atualizar a página.
# Altere este texto se quiser invalidar todos os logins salvos.
LOGIN_SECRET = "EIROX_DRE_ENTERPRISE_PREMIUM_v2_2_031730"
LOGIN_DURACAO_SEGUNDOS = 12 * 60 * 60  # 12 horas

# =========================================================
# ESTRUTURA DO DRE
# =========================================================
DRE_TEMPLATE = [
    (10, "DRE", "1. RECEITA OPERACIONAL BRUTA", 0, "subtotal_azul"),
    (20, "DRE", "Receita de Vendas de Mercadorias", 1, "detalhe"),
    (30, "DRE", "2. (-) DEDUÇÕES DA RECEITA BRUTA", 0, "subtotal_azul"),
    (40, "DRE", "ICMS", 1, "detalhe"),
    (50, "DRE", "DAS", 1, "detalhe"),
    (60, "DRE", "ISSQN", 1, "detalhe"),
    (70, "DRE", "3. (=) RECEITA OPERACIONAL LÍQUIDA", 0, "resultado_amarelo"),
    (80, "DRE", "4. (-) CUSTOS DAS VENDAS", 0, "subtotal_azul"),
    (90, "DRE", "Mercadorias (CMP)", 1, "detalhe"),
    (100, "DRE", "5. (=) LUCRO BRUTO", 0, "resultado_amarelo"),
    (110, "DRE", "6. (-) DESPESAS OPERACIONAIS", 0, "subtotal_azul"),
    (120, "DRE", "Despesas com Pessoal", 1, "agrupador"),
    (130, "DRE", "Alimentação", 2, "detalhe"),
    (140, "DRE", "Pagamento de Férias", 2, "detalhe"),
    (150, "DRE", "FGTS", 2, "detalhe"),
    (160, "DRE", "Salários Fixos + Horas Extras", 2, "detalhe"),
    (170, "DRE", "Comissões e Premiações", 2, "detalhe"),
    (180, "DRE", "INSS", 2, "detalhe"),
    (190, "DRE", "IRRF Folha", 2, "detalhe"),
    (200, "DRE", "Cursos, Treinamentos, Viagens", 2, "detalhe"),
    (210, "DRE", "PPRA<PCMSO e Exames", 2, "detalhe"),
    (220, "DRE", "Rescisão (guia)", 2, "detalhe"),
    (230, "DRE", "Provisão 13 e Férias", 2, "detalhe"),
    (240, "DRE", "Uniformes", 2, "detalhe"),
    (250, "DRE", "Vale Transporte", 2, "detalhe"),
    (260, "DRE", "Pro-Labore", 2, "detalhe"),
    (270, "DRE", "Outras Despesas com Pessoal", 2, "detalhe"),
    (280, "DRE", "Despesas Administrativas e Ocupação", 1, "agrupador"),
    (290, "DRE", "Aluguel&IPTU", 2, "detalhe"),
    (300, "DRE", "Seguro Imóvel", 2, "detalhe"),
    (310, "DRE", "Agua/Luz/Fone/Net", 2, "detalhe"),
    (320, "DRE", "Combustível", 2, "detalhe"),
    (330, "DRE", "Manutenção em Geral", 2, "detalhe"),
    (340, "DRE", "Manutenção Veículos/Motos", 2, "detalhe"),
    (350, "DRE", "Mat. de Escritório/Informática", 2, "detalhe"),
    (360, "DRE", "Mat. De Limpeza", 2, "detalhe"),
    (370, "DRE", "Viagens", 2, "detalhe"),
    (380, "DRE", "Contábil (Terceiros)", 2, "detalhe"),
    (390, "DRE", "Sistemas (Terceiros)", 2, "detalhe"),
    (400, "DRE", "Juridico (Terceiros)", 2, "detalhe"),
    (410, "DRE", "Assessorias/Consultorias/Treinamentos", 2, "detalhe"),
    (420, "DRE", "Taxas, Licenças e Contrib.", 2, "detalhe"),
    (430, "DRE", "Outros (Terceiros)", 2, "detalhe"),
    (440, "DRE", "Outros (Despesas)", 2, "detalhe"),
    (450, "DRE", "Quebra de caixa", 2, "detalhe"),
    (460, "DRE", "Despesas com Vendas e Marketing", 1, "agrupador"),
    (470, "DRE", "Marketing/Publicidade", 2, "detalhe"),
    (480, "DRE", "Sistema Fidelidade", 2, "detalhe"),
    (490, "DRE", "Frete", 2, "detalhe"),
    (500, "DRE", "Associação de Classe (Royalties)", 2, "detalhe"),
    (510, "DRE", "7. (=) RESULTADO ANTES DO RESULTADO FINANCEIRO (EBITDA/LAJIDA)", 0, "resultado_amarelo"),
    (520, "DRE", "8. (+/-) RESULTADO FINANCEIRO LÍQUIDO", 0, "subtotal_azul"),
    (530, "DRE", "Juros Boletos", 1, "detalhe"),
    (540, "DRE", "Tarifas Bancarias", 1, "detalhe"),
    (550, "DRE", "Taxas de Cartão (MDR)", 1, "detalhe"),
    (560, "DRE", "9. (=) RESULTADO ANTES DOS TRIBUTOS (LAIR)", 0, "resultado_amarelo"),
    (570, "DRE", "10. (-) TRIBUTOS SOBRE O LUCRO", 0, "subtotal_azul"),
    (580, "DRE", "IRPJ/CSLL", 1, "detalhe"),
    (590, "DRE", "11. (=) LUCRO LÍQUIDO DO EXERCÍCIO", 0, "resultado_amarelo"),
    (600, "FLUXO", "--- CONCILIAÇÃO DE FLUXO DE CAIXA (SAÍDAS NÃO-DRE) ---", 0, "subtotal_azul"),
    (610, "FLUXO", "Investimentos - Reformas", 1, "detalhe"),
    (620, "FLUXO", "Investimentos - Equipamentos", 1, "detalhe"),
    (630, "FLUXO", "Investimentos - Expansão", 1, "detalhe"),
    (640, "FLUXO", "Pagamento de Empréstimos (Principal)", 1, "detalhe"),
    (650, "FLUXO", "Retiradas (Distribuição de Lucros)", 1, "detalhe"),
    (660, "FLUXO", "Parcelamento de Impostos", 1, "detalhe"),
    (670, "FLUXO", "Divergência Saídas", 1, "detalhe"),
    (680, "FLUXO", "Posição Final", 0, "resultado_amarelo"),
    (690, "INDICADORES", "Receitas", 0, "detalhe"),
    (700, "INDICADORES", "Estoque a Custo", 0, "detalhe"),
    (710, "INDICADORES", "CMV", 0, "detalhe"),
    (720, "PONTO DE EQUILÍBRIO", "Receita Total", 0, "detalhe"),
    (730, "PONTO DE EQUILÍBRIO", "Despesas Fixas", 0, "detalhe"),
    (740, "PONTO DE EQUILÍBRIO", "Despesas Variáveis", 0, "detalhe"),
    (750, "PONTO DE EQUILÍBRIO", "Custos e Despesas Fixas Totais", 0, "detalhe"),
    (760, "PONTO DE EQUILÍBRIO", "Custo Médio de Venda (CMV)", 0, "detalhe"),
    (770, "PONTO DE EQUILÍBRIO", "Total de Custos Variáveis (CMV+Despesas Variáveis)", 0, "detalhe"),
    (780, "PONTO DE EQUILÍBRIO", "Margem de Contribuição Total", 0, "resultado_amarelo"),
    (790, "PONTO DE EQUILÍBRIO", "Margem de Contribuição Percentual", 0, "resultado_amarelo"),
    (800, "PONTO DE EQUILÍBRIO", "Ponto de Equilíbrio em Valor Monetário (Receita)", 0, "resultado_amarelo"),
]

# =========================================================
# CSS PREMIUM
# =========================================================
st.markdown("""
<style>
:root{
    --bg:#07111f; --panel:#0b1728; --panel2:#101c2f; --line:#20334d; --text:#f5f7fb; --muted:#9aa7b8;
    --blue:#00a8ff; --blue2:#0068ff; --yellow:#ffd21f; --green:#32e875; --red:#ff5470;
}
html, body, [data-testid="stAppViewContainer"]{background: radial-gradient(circle at top, #0c2038 0%, #06101d 35%, #030914 100%) !important; color:var(--text)!important;}
[data-testid="stSidebar"]{background:linear-gradient(180deg,#050b14 0%,#07111f 100%)!important; border-right:1px solid #103a5c;}
[data-testid="stSidebar"] *{color:#f5f7fb!important;}
.block-container{padding-top:1.2rem; padding-bottom:2rem; max-width: 1600px;}
.sidebar-logo{display:flex; justify-content:center; align-items:center; margin:12px 0 8px 0;}
.sidebar-title{text-align:center; font-size:1.05rem; font-weight:800; margin-bottom:2px; color:#ffffff;}
.sidebar-subtitle{text-align:center; color:#6cc7ff!important; font-size:.82rem; margin-bottom:18px;}
.sidebar-section{font-size:.78rem; font-weight:800; letter-spacing:.12em; color:#7ebeff!important; margin:18px 0 8px 0; text-transform:uppercase;}
div[data-baseweb="select"] > div{background:#0d1b2e!important; border-color:#1f4d77!important; border-radius:14px!important;}
.stMultiSelect [data-baseweb="tag"]{background:#00a8ff!important; color:white!important; border-radius:10px!important; font-weight:700!important;}
.hero{padding:24px 26px; border-radius:26px; background:linear-gradient(135deg,rgba(0,168,255,.18),rgba(8,18,32,.94)); border:1px solid rgba(0,168,255,.28); box-shadow: 0 12px 35px rgba(0,0,0,.30); margin-bottom:18px;}
.hero h1{font-size:3.15rem; line-height:1; margin:0; font-weight:900; letter-spacing:-.04em; color:#fff;}
.hero p{font-size:1.05rem; margin:.65rem 0 0 0; color:#8bd4ff;}
.small-meta{font-size:.86rem; color:#aab7c8; margin-top:10px;}
.kpi-card{background:linear-gradient(180deg,#10223a,#091728); border:1px solid rgba(0,168,255,.28); border-radius:22px; padding:20px 18px; box-shadow: 0 8px 28px rgba(0,0,0,.22); min-height:126px;}
.kpi-card .label{font-size:.82rem; text-transform:uppercase; letter-spacing:.09em; color:#9fb5cc; font-weight:800; margin-bottom:8px;}
.kpi-card .value{font-size:1.55rem; color:#fff; font-weight:900; white-space:nowrap;}
.kpi-card .delta{font-size:.86rem; margin-top:8px; color:#9fb5cc; font-weight:700;}
.delta-pos{color:#32e875!important;} .delta-neg{color:#ff5470!important;}
.section-title{font-size:1.55rem; font-weight:900; color:#fff; margin:26px 0 10px 0;}
.section-caption{color:#9aa7b8; margin:-4px 0 16px 0;}
.eirox-table-wrap{overflow:auto; border:1px solid rgba(0,168,255,.25); border-radius:20px; max-height:720px; box-shadow:0 12px 35px rgba(0,0,0,.28); background:#091526;}
table.eirox-table{border-collapse:collapse; width:max-content; min-width:100%; font-size:.92rem;}
table.eirox-table th{position:sticky; top:0; z-index:2; background:#141b27; color:#c9d3df; padding:13px 14px; border-bottom:1px solid #2a3d55; border-right:1px solid #2a3d55; text-align:right; white-space:nowrap;}
table.eirox-table th:first-child{left:0; z-index:3; text-align:left; min-width:440px;}
table.eirox-table td{padding:12px 14px; border-bottom:1px solid #20334d; border-right:1px solid #20334d; text-align:right; white-space:nowrap; color:#fff; font-weight:650;}
table.eirox-table td:first-child{position:sticky; left:0; z-index:1; background:#0b1728; text-align:left; min-width:440px; max-width:520px; white-space:normal; font-weight:700;}
table.eirox-table tr.detalhe td{background:#0b1728;}
table.eirox-table tr.subtotal_azul td{background:#cfe3f8!important; color:#00101e!important; font-weight:950;}
table.eirox-table tr.subtotal_azul td:first-child{background:#cfe3f8!important; color:#00101e!important;}
table.eirox-table tr.resultado_amarelo td{background:#ffd21f!important; color:#030914!important; font-weight:950;}
table.eirox-table tr.resultado_amarelo td:first-child{background:#ffd21f!important; color:#030914!important;}
table.eirox-table tr.agrupador td, table.eirox-table tr.grupo_italico td{background:#1b2534!important; color:#fff!important; font-weight:900; font-style:italic;}
table.eirox-table tr.agrupador td:first-child, table.eirox-table tr.grupo_italico td:first-child{background:#1b2534!important;}
.audit-box{background:rgba(255,84,112,.12); border:1px solid rgba(255,84,112,.35); padding:16px 18px; border-radius:18px; color:#ffb5c1; font-weight:800;}
.ok-box{background:rgba(50,232,117,.12); border:1px solid rgba(50,232,117,.35); padding:16px 18px; border-radius:18px; color:#baf7cb; font-weight:800;}
.footer{color:#6d7d90; font-size:.78rem; text-align:center; margin-top:30px; padding-top:18px; border-top:1px solid rgba(255,255,255,.08);}
div[role="radiogroup"] label{background:linear-gradient(90deg,rgba(0,168,255,.12),rgba(255,255,255,.03)); border:1px solid rgba(0,168,255,.20); border-radius:14px; padding:10px 12px!important; margin:7px 0!important; transition:all .18s ease-in-out;}
div[role="radiogroup"] label:hover{border-color:rgba(0,168,255,.55); background:linear-gradient(90deg,rgba(0,168,255,.25),rgba(255,255,255,.05));}
.login-card{max-width:460px; margin:8vh auto 0 auto; padding:34px 32px; border-radius:26px; background:linear-gradient(180deg,#10223a,#07111f); border:1px solid rgba(0,168,255,.32); box-shadow:0 18px 50px rgba(0,0,0,.35);}
.login-title{text-align:center; font-size:2rem; font-weight:950; color:#fff; margin:12px 0 6px 0;}
.login-sub{text-align:center; color:#8bd4ff; font-weight:700; margin-bottom:20px;}
</style>
""", unsafe_allow_html=True)

# =========================================================
# FUNÇÕES UTILITÁRIAS
# =========================================================
def normalizar_texto(txt) -> str:
    txt = "" if pd.isna(txt) else str(txt)
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]+", " ", txt.upper()).strip()

def encontrar_arquivo(paths: list[Path]) -> Path | None:
    for p in paths:
        if p.exists():
            return p
    return None

def img_base64(path: Path) -> str | None:
    if not path or not path.exists():
        return None
    return base64.b64encode(path.read_bytes()).decode("utf-8")

def brl(v) -> str:
    try:
        v = float(v)
    except Exception:
        v = 0.0
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def pct(v) -> str:
    try:
        v = float(v)
    except Exception:
        v = 0.0
    if abs(v) <= 1.5:
        v = v * 100
    return f"{v:.2f}%".replace(".", ",")

def parse_float(v) -> float:
    if pd.isna(v):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace("%", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def mes_sort_key(mes: str) -> tuple[int, int]:
    s = str(mes).lower().strip()
    if s in ("sem mês", "sem mes", "nan", "none", ""):
        return (9999, 99)
    m = re.match(r"([a-zç]{3})/(\d{2})", s)
    if not m:
        return (9999, 98)
    return (2000 + int(m.group(2)), MESES_ORDEM.get(m.group(1)[:3], 99))

def month_label_from_date(v) -> str:
    dt = pd.to_datetime(v, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return "Sem mês"
    return f"{MES_NUM_TO_ABR[int(dt.month)]}/{str(int(dt.year))[-2:]}"

def month_label_from_anomes(v) -> str:
    s = str(v)
    m = re.search(r"(20\d{2})[-/](\d{1,2})", s)
    if m:
        ano = int(m.group(1))
        mes = int(m.group(2))
        return f"{MES_NUM_TO_ABR.get(mes, 'sem')}/{str(ano)[-2:]}"
    return month_label_from_date(v)

def meses_fechados_ano_atual(meses: list[str]) -> list[str]:
    hoje = datetime.today()
    ano_atual, mes_atual = hoje.year, hoje.month
    filtrados = [m for m in meses if mes_sort_key(m)[0] == ano_atual and mes_sort_key(m)[1] < mes_atual]
    if filtrados:
        return filtrados
    anos = sorted({mes_sort_key(m)[0] for m in meses if mes_sort_key(m)[0] < 9999})
    if not anos:
        return meses
    ano_base = anos[-1]
    return [m for m in meses if mes_sort_key(m)[0] == ano_base and mes_sort_key(m)[1] < 99]

def localizar_pasta(nome_com_acento: str, nome_sem_acento: str) -> Path | None:
    candidatos = [
        APP_DIR / nome_com_acento,
        APP_DIR / nome_sem_acento,
        APP_DIR.parent / nome_com_acento,
        APP_DIR.parent / nome_sem_acento,
        Path.cwd() / nome_com_acento,
        Path.cwd() / nome_sem_acento,
        Path.home() / "Desktop" / "Dre" / nome_com_acento,
        Path.home() / "Desktop" / "Dre" / nome_sem_acento,
    ]
    for p in candidatos:
        if p.exists() and p.is_dir():
            return p
    return None

def arquivos_excel(pasta: Path | None) -> list[Path]:
    if not pasta:
        return []
    return sorted([p for p in pasta.rglob("*.xls*") if not p.name.startswith("~$")])

def mes_arquivo_para_dre(path: Path) -> str | None:
    nome = normalizar_texto(path.stem)
    ano_match = re.search(r"20(\d{2})", nome)
    ano2 = ano_match.group(1) if ano_match else "26"
    mapa = {
        "JANEIRO": "jan", "JAN": "jan", "FEVEREIRO": "fev", "FEV": "fev", "MARCO": "mar", "MAR": "mar",
        "ABRIL": "abr", "ABR": "abr", "MAIO": "mai", "MAI": "mai", "JUNHO": "jun", "JUN": "jun",
        "JULHO": "jul", "JUL": "jul", "AGOSTO": "ago", "AGO": "ago", "SETEMBRO": "set", "SET": "set",
        "OUTUBRO": "out", "OUT": "out", "NOVEMBRO": "nov", "NOV": "nov", "DEZEMBRO": "dez", "DEZ": "dez",
    }
    for chave, abrev in mapa.items():
        if chave in nome:
            return f"{abrev}/{ano2}"
    return None

# =========================================================
# LEITURA DINÂMICA DAS PASTAS
# =========================================================
def read_all_sheets(path: Path) -> list[pd.DataFrame]:
    out = []
    try:
        xls = pd.ExcelFile(path)
        for sh in xls.sheet_names:
            try:
                df = pd.read_excel(path, sheet_name=sh)
                if not df.empty:
                    df["Arquivo Origem"] = path.name
                    out.append(df)
            except Exception:
                pass
    except Exception:
        pass
    return out

def carregar_plano_contas(pasta: Path | None) -> pd.DataFrame:
    frames = []
    for arq in arquivos_excel(pasta):
        for df in read_all_sheets(arq):
            cols = [str(c).strip() for c in df.columns]
            df.columns = cols
            if "Plano de contas" in cols or "Plano de Contas" in cols:
                frames.append(df)
    if not frames:
        return pd.DataFrame()
    plano = pd.concat(frames, ignore_index=True)
    plano.columns = [str(c).strip() for c in plano.columns]
    col_plano = "Plano de contas" if "Plano de contas" in plano.columns else "Plano de Contas"
    if "Destino na DRE Estruturada" not in plano.columns:
        plano["Destino na DRE Estruturada"] = ""
    if "Subgrupo" not in plano.columns:
        plano["Subgrupo"] = ""
    if "Grupo" not in plano.columns:
        plano["Grupo"] = ""
    plano["Plano_Normalizado"] = plano[col_plano].apply(normalizar_texto)
    plano["Destino DRE Legenda"] = plano["Destino na DRE Estruturada"].fillna("").astype(str).str.strip()
    plano["Subgrupo Legenda"] = plano["Subgrupo"].fillna("").astype(str).str.strip()
    plano["Grupo Legenda"] = plano["Grupo"].fillna("").astype(str).str.strip()
    return plano.drop_duplicates("Plano_Normalizado")

def carregar_vendas(pasta: Path | None) -> pd.DataFrame:
    frames = []
    for arq in arquivos_excel(pasta):
        for df in read_all_sheets(arq):
            cols_norm = {normalizar_texto(c): c for c in df.columns}
            if "ANO MES" in cols_norm or "ANO MES" in cols_norm or "ANO MÊS" in cols_norm:
                frames.append(df)
            elif "FORMA DE PAGAMENTO" in cols_norm and ("TOTAL" in cols_norm or "VALOR" in cols_norm):
                frames.append(df)
    if not frames:
        return pd.DataFrame(columns=["Mês", "Loja", "Valor Receita", "Forma de Pagamento", "Arquivo Origem"])
    vendas = pd.concat(frames, ignore_index=True)
    vendas.columns = [str(c).strip() for c in vendas.columns]
    col_mes = next((c for c in vendas.columns if normalizar_texto(c) in ["ANO MES", "DATA", "MES", "MÊS"]), None)
    col_val = "Total" if "Total" in vendas.columns else ("Valor" if "Valor" in vendas.columns else None)
    col_loja = next((c for c in ["Un. Neg.", "unidade", "Unidade", "Loja"] if c in vendas.columns), None)
    col_forma = next((c for c in vendas.columns if "FORMA" in normalizar_texto(c)), None)
    vendas["Mês"] = vendas[col_mes].apply(month_label_from_anomes) if col_mes else "Sem mês"
    vendas["Loja"] = vendas[col_loja] if col_loja else ""
    vendas["Valor Receita"] = vendas[col_val].apply(parse_float) if col_val else 0.0
    vendas["Forma de Pagamento"] = vendas[col_forma] if col_forma else ""
    return vendas[["Mês", "Loja", "Valor Receita", "Forma de Pagamento", "Arquivo Origem"]]

def carregar_contas(pasta: Path | None, plano: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    frames = []
    for arq in arquivos_excel(pasta):
        for df in read_all_sheets(arq):
            cols = [str(c).strip() for c in df.columns]
            df.columns = cols
            if "Valor Documento" in cols and "Plano de Contas" in cols:
                frames.append(df)
    if not frames:
        return pd.DataFrame(), pd.DataFrame()
    contas = pd.concat(frames, ignore_index=True)
    contas.columns = [str(c).strip() for c in contas.columns]

    col_data = "Data Pagamento" if "Data Pagamento" in contas.columns else ("Data Emissão" if "Data Emissão" in contas.columns else None)
    col_loja = "Unidade" if "Unidade" in contas.columns else ("Loja" if "Loja" in contas.columns else None)
    contas["Mês"] = contas[col_data].apply(month_label_from_date) if col_data else "Sem mês"
    contas["Loja"] = contas[col_loja] if col_loja else ""
    contas["Valor Documento"] = contas["Valor Documento"].apply(parse_float)
    contas["Plano_Normalizado"] = contas["Plano de Contas"].apply(normalizar_texto)

    if not plano.empty:
        contas = contas.merge(
            plano[["Plano_Normalizado", "Destino DRE Legenda", "Subgrupo Legenda", "Grupo Legenda"]],
            on="Plano_Normalizado",
            how="left",
        )
    else:
        contas["Destino DRE Legenda"] = ""
        contas["Subgrupo Legenda"] = ""
        contas["Grupo Legenda"] = ""

    contas["Destino DRE"] = contas["Destino DRE Legenda"].fillna("").astype(str).str.strip()
    contas.loc[contas["Destino DRE"] == "", "Destino DRE"] = "NÃO CLASSIFICADO"

    # Algumas despesas fiscais costumam chegar sem plano; tenta classificar por texto para reduzir não classificados.
    texto_busca = (
        contas.get("Plano de Contas", "").astype(str) + " " +
        contas.get("Credor", "").astype(str) + " " +
        contas.get("Descrição", "").astype(str)
    ).apply(normalizar_texto)
    contas.loc[(contas["Destino DRE"] == "NÃO CLASSIFICADO") & texto_busca.str.contains("ICMS", na=False), "Destino DRE"] = "ICMS"
    contas.loc[(contas["Destino DRE"] == "NÃO CLASSIFICADO") & texto_busca.str.contains("DAS|SIMPLES", na=False, regex=True), "Destino DRE"] = "DAS"
    contas.loc[(contas["Destino DRE"] == "NÃO CLASSIFICADO") & texto_busca.str.contains("ISS", na=False), "Destino DRE"] = "ISSQN"

    contas_class = contas[contas["Destino DRE"] != "NÃO CLASSIFICADO"].copy()
    nao = (
        contas[contas["Destino DRE"] == "NÃO CLASSIFICADO"]
        .groupby(["Plano de Contas", "Mês"], dropna=False)["Valor Documento"]
        .agg(["sum", "count"])
        .reset_index()
        .rename(columns={"sum": "Valor", "count": "Qtde"})
    )
    return contas_class, nao

def escolher_coluna_estoque(df: pd.DataFrame) -> str | None:
    normalizadas = {col: normalizar_texto(col) for col in df.columns}
    prioridades = [
        "ESTOQUE X CUSTO MEDIO", "ESTOQUE A CUSTO", "ESTOQUE CUSTO MEDIO",
        "VALOR ESTOQUE", "VALOR DO ESTOQUE", "CUSTO TOTAL", "TOTAL CUSTO"
    ]
    for alvo in prioridades:
        for col, ncol in normalizadas.items():
            if alvo in ncol:
                return col
    for col, ncol in normalizadas.items():
        if "ESTOQUE" in ncol and ("CUSTO" in ncol or "VALOR" in ncol or "TOTAL" in ncol):
            if not any(x in ncol for x in ["QTD", "QTDE", "QUANTIDADE", "COD", "EAN"]):
                return col
    return None

def carregar_estoque(pasta: Path | None) -> pd.DataFrame:
    rows = []
    for arq in arquivos_excel(pasta):
        mes = mes_arquivo_para_dre(arq)
        if not mes:
            continue
        try:
            xls = pd.ExcelFile(arq)
            for aba in xls.sheet_names:
                df = pd.read_excel(arq, sheet_name=aba)
                if df.empty:
                    continue
                df.columns = [str(c).strip() for c in df.columns]
                col_val = escolher_coluna_estoque(df)
                if not col_val:
                    continue
                col_loja = next((c for c in ["Un. Neg.", "Unidade", "Loja", "unidade"] if c in df.columns), None)
                # Se houver linha Total, usa somente ela para evitar duplicidade.
                if col_loja:
                    total_mask = df[col_loja].astype(str).apply(normalizar_texto).str.contains("TOTAL", na=False)
                    if total_mask.any():
                        valor = df.loc[total_mask, col_val].apply(parse_float).sum()
                        rows.append({"Mês": mes, "Loja": "Total", "Valor Estoque": valor, "Arquivo Origem": arq.name})
                        break
                    else:
                        for _, r in df.iterrows():
                            rows.append({"Mês": mes, "Loja": r.get(col_loja, ""), "Valor Estoque": parse_float(r.get(col_val, 0)), "Arquivo Origem": arq.name})
                        break
                else:
                    valor = df[col_val].apply(parse_float).sum()
                    rows.append({"Mês": mes, "Loja": "Total", "Valor Estoque": valor, "Arquivo Origem": arq.name})
                    break
        except Exception:
            continue
    if not rows:
        return pd.DataFrame(columns=["Mês", "Loja", "Valor Estoque", "Arquivo Origem"])
    estoque = pd.DataFrame(rows)
    # Se houver linha Total para o mês, mantém Total e descarta lojas individuais daquele mês.
    finais = []
    for mes, g in estoque.groupby("Mês"):
        gt = g[g["Loja"].astype(str).apply(normalizar_texto).str.contains("TOTAL", na=False)]
        if not gt.empty:
            finais.append(gt.sort_values("Arquivo Origem").tail(1))
        else:
            finais.append(g)
    return pd.concat(finais, ignore_index=True)

# =========================================================
# CÁLCULO DINÂMICO DO DRE
# =========================================================
def build_dynamic_dre() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    pasta_contas = localizar_pasta("CONTAS A PAGAR - DRE", "CONTAS A PAGAR - DRE")
    pasta_plano = localizar_pasta("PLANO DE CONTAS - LEGENDA", "PLANO DE CONTAS - LEGENDA")
    pasta_estoque = localizar_pasta("POSIÇÃO DE ESTOQUE", "POSICAO DE ESTOQUE")
    pasta_venda = localizar_pasta("VENDA POR PAGAMENTO", "VENDA POR PAGAMENTO")

    status_pastas = {
        "CONTAS A PAGAR - DRE": pasta_contas,
        "PLANO DE CONTAS - LEGENDA": pasta_plano,
        "POSIÇÃO DE ESTOQUE": pasta_estoque,
        "VENDA POR PAGAMENTO": pasta_venda,
    }

    if not all(status_pastas.values()):
        raise FileNotFoundError("Pastas de base não encontradas. Usando fallback do consolidado.")

    plano = carregar_plano_contas(pasta_plano)
    vendas = carregar_vendas(pasta_venda)
    contas_class, nao_class = carregar_contas(pasta_contas, plano)
    estoque = carregar_estoque(pasta_estoque)

    meses = sorted(set(vendas["Mês"].unique()) | set(contas_class["Mês"].unique()) | set(estoque["Mês"].unique()), key=mes_sort_key)
    meses = [m for m in meses if m not in ["Sem mês", "nan", "None", ""]]

    receita_mes = vendas.groupby("Mês")["Valor Receita"].sum().to_dict()
    estoque_mes = estoque.groupby("Mês")["Valor Estoque"].sum().to_dict()
    detalhe_mes = contas_class.groupby(["Mês", "Destino DRE"])["Valor Documento"].sum().to_dict()
    subgrupo_mes = contas_class.groupby(["Mês", "Subgrupo Legenda"])["Valor Documento"].sum().to_dict()

    detail_lines = {linha for _, _, linha, nivel, tipo in DRE_TEMPLATE if tipo == "detalhe"}
    def get_detail(mes: str, linha: str) -> float:
        if linha == "Receita de Vendas de Mercadorias":
            return float(receita_mes.get(mes, 0))
        if linha == "Receitas" or linha == "Receita Total":
            return float(receita_mes.get(mes, 0))
        if linha == "Estoque a Custo":
            return float(estoque_mes.get(mes, 0))
        if linha == "CMV" or linha == "Custo Médio de Venda (CMV)":
            return float(detalhe_mes.get((mes, "Mercadorias (CMP)"), 0))
        return float(detalhe_mes.get((mes, linha), 0))

    template_df = pd.DataFrame(DRE_TEMPLATE, columns=["Ordem", "Seção", "Linha DRE", "Nível", "Tipo"])
    results = []
    for mes in meses:
        valores = {}
        # detalhes básicos
        for _, row in template_df.iterrows():
            linha = row["Linha DRE"]
            if row["Tipo"] == "detalhe":
                valores[linha] = get_detail(mes, linha)

        # agrupadores por subgrupo
        for grupo in ["Despesas com Pessoal", "Despesas Administrativas e Ocupação", "Despesas com Vendas e Marketing"]:
            valores[grupo] = float(subgrupo_mes.get((mes, grupo), 0))
            # fallback: soma detalhes do intervalo quando legenda não preenche subgrupo
            if valores[grupo] == 0:
                idx = template_df.index[template_df["Linha DRE"] == grupo].tolist()[0]
                linhas = []
                for j in range(idx + 1, len(template_df)):
                    if template_df.loc[j, "Nível"] <= 1:
                        break
                    if template_df.loc[j, "Tipo"] == "detalhe":
                        linhas.append(template_df.loc[j, "Linha DRE"])
                valores[grupo] = sum(valores.get(x, 0) for x in linhas)

        receita = valores.get("Receita de Vendas de Mercadorias", 0)
        deducoes = sum(valores.get(x, 0) for x in ["ICMS", "DAS", "ISSQN"])
        receita_liq = receita - deducoes
        cmv = valores.get("Mercadorias (CMP)", 0)
        lucro_bruto = receita_liq - cmv
        despesas_op = sum(valores.get(x, 0) for x in ["Despesas com Pessoal", "Despesas Administrativas e Ocupação", "Despesas com Vendas e Marketing"])
        ebitda = lucro_bruto - despesas_op
        financeiro = sum(valores.get(x, 0) for x in ["Juros Boletos", "Tarifas Bancarias", "Taxas de Cartão (MDR)"])
        lair = ebitda + financeiro
        tributos = valores.get("IRPJ/CSLL", 0)
        lucro_liq = lair - tributos
        fluxo = sum(valores.get(x, 0) for x in [
            "Investimentos - Reformas", "Investimentos - Equipamentos", "Investimentos - Expansão",
            "Pagamento de Empréstimos (Principal)", "Retiradas (Distribuição de Lucros)",
            "Parcelamento de Impostos", "Divergência Saídas"
        ])
        pos_final = lucro_liq - fluxo

        desp_variaveis = valores.get("Despesas com Vendas e Marketing", 0)
        desp_fixas = despesas_op - desp_variaveis
        total_var = cmv + desp_variaveis
        margem_cont = receita - total_var
        margem_pct = (margem_cont / receita) if receita else 0
        ponto_eq = (desp_fixas / margem_pct) if margem_pct else 0

        formulas = {
            "1. RECEITA OPERACIONAL BRUTA": receita,
            "2. (-) DEDUÇÕES DA RECEITA BRUTA": deducoes,
            "3. (=) RECEITA OPERACIONAL LÍQUIDA": receita_liq,
            "4. (-) CUSTOS DAS VENDAS": cmv,
            "5. (=) LUCRO BRUTO": lucro_bruto,
            "6. (-) DESPESAS OPERACIONAIS": despesas_op,
            "7. (=) RESULTADO ANTES DO RESULTADO FINANCEIRO (EBITDA/LAJIDA)": ebitda,
            "8. (+/-) RESULTADO FINANCEIRO LÍQUIDO": financeiro,
            "9. (=) RESULTADO ANTES DOS TRIBUTOS (LAIR)": lair,
            "10. (-) TRIBUTOS SOBRE O LUCRO": tributos,
            "11. (=) LUCRO LÍQUIDO DO EXERCÍCIO": lucro_liq,
            "--- CONCILIAÇÃO DE FLUXO DE CAIXA (SAÍDAS NÃO-DRE) ---": fluxo,
            "Posição Final": pos_final,
            "Receitas": receita,
            "Estoque a Custo": float(estoque_mes.get(mes, 0)),
            "CMV": cmv,
            "Receita Total": receita,
            "Despesas Fixas": desp_fixas,
            "Despesas Variáveis": desp_variaveis,
            "Custos e Despesas Fixas Totais": desp_fixas,
            "Custo Médio de Venda (CMV)": cmv,
            "Total de Custos Variáveis (CMV+Despesas Variáveis)": total_var,
            "Margem de Contribuição Total": margem_cont,
            "Margem de Contribuição Percentual": margem_pct,
            "Ponto de Equilíbrio em Valor Monetário (Receita)": ponto_eq,
        }
        valores.update(formulas)

        for _, row in template_df.iterrows():
            linha = row["Linha DRE"]
            valor = float(valores.get(linha, 0))
            if linha == "Margem de Contribuição Percentual":
                perc = valor
            elif receita:
                perc = valor / receita
            else:
                perc = 0.0
            results.append({
                "Ordem": row["Ordem"],
                "Seção": row["Seção"],
                "Linha DRE": linha,
                "Nível": row["Nível"],
                "Tipo": row["Tipo"],
                "Mês": mes,
                "Valor": valor,
                "% Receita": perc,
            })

    dados = pd.DataFrame(results)
    if not vendas.empty:
        receita_loja = vendas.groupby(["Loja", "Mês"], dropna=False)["Valor Receita"].sum().reset_index().rename(columns={"Valor Receita": "Receita"})
    else:
        receita_loja = pd.DataFrame(columns=["Loja", "Mês", "Receita"])
    if not contas_class.empty:
        desp_loja = contas_class.groupby(["Loja", "Mês"], dropna=False)["Valor Documento"].sum().reset_index().rename(columns={"Valor Documento": "Despesas"})
    else:
        desp_loja = pd.DataFrame(columns=["Loja", "Mês", "Despesas"])
    resumo_loja = receita_loja.merge(desp_loja, on=["Loja", "Mês"], how="outer").fillna(0)
    resumo_loja["Resultado Caixa Simplificado"] = resumo_loja["Receita"] - resumo_loja["Despesas"]

    checks = pd.DataFrame([
        {"Checagem": "Modo de leitura", "Valor": "DINÂMICO PELAS PASTAS"},
        {"Checagem": "Arquivos de vendas lidos", "Valor": len(arquivos_excel(pasta_venda))},
        {"Checagem": "Arquivos de contas lidos", "Valor": len(arquivos_excel(pasta_contas))},
        {"Checagem": "Arquivos de estoque lidos", "Valor": len(arquivos_excel(pasta_estoque))},
        {"Checagem": "Arquivos de plano lidos", "Valor": len(arquivos_excel(pasta_plano))},
        {"Checagem": "Última atualização", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
    ])

    return dados, resumo_loja, nao_class, checks, status_pastas

def carregar_fallback_consolidado() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    base_path = encontrar_arquivo(POSSIVEIS_BASES)
    if not base_path:
        raise FileNotFoundError("Não encontrei as pastas nem o arquivo DRE_Consolidado_Moderno.xlsx.")
    sheets = pd.ExcelFile(base_path).sheet_names
    if "DADOS_DRE" not in sheets:
        raise ValueError("A aba DADOS_DRE não foi encontrada no consolidado.")
    dados = pd.read_excel(base_path, sheet_name="DADOS_DRE")
    resumo_loja = pd.read_excel(base_path, sheet_name="RESUMO_LOJA") if "RESUMO_LOJA" in sheets else pd.DataFrame()
    nao = pd.read_excel(base_path, sheet_name="NAO_CLASSIFICADOS") if "NAO_CLASSIFICADOS" in sheets else pd.DataFrame()
    checks = pd.DataFrame([{"Checagem": "Modo de leitura", "Valor": "FALLBACK CONSOLIDADO"}, {"Checagem": "Arquivo", "Valor": str(base_path)}])
    return dados, resumo_loja, nao, checks, {}

def carregar_base_dre():
    try:
        dados, resumo_loja, nao, checks, status = build_dynamic_dre()
    except Exception as e:
        dados, resumo_loja, nao, checks, status = carregar_fallback_consolidado()
        checks = pd.concat([checks, pd.DataFrame([{"Checagem": "Aviso", "Valor": f"Fallback usado: {e}"}])], ignore_index=True)

    dados.columns = [str(c).strip() for c in dados.columns]
    dados["Valor"] = dados["Valor"].apply(parse_float)
    dados["% Receita"] = dados["% Receita"].apply(parse_float)
    dados["Mês"] = dados["Mês"].astype(str).str.strip()
    dados["Linha DRE"] = dados["Linha DRE"].astype(str).str.strip()
    dados["Tipo"] = dados["Tipo"].fillna("detalhe").astype(str).str.strip()
    dados["Ordem"] = pd.to_numeric(dados["Ordem"], errors="coerce").fillna(999999).astype(int)
    dados["Nível"] = pd.to_numeric(dados["Nível"], errors="coerce").fillna(0).astype(int)
    return dados, resumo_loja, nao, checks, status

# =========================================================
# FUNÇÕES DO DASH
# =========================================================
def valor_linha(df: pd.DataFrame, linha_contains: str, mes: str) -> float:
    filtro = df["Linha DRE"].str.upper().str.contains(linha_contains.upper(), regex=False, na=False) & (df["Mês"] == mes)
    if not filtro.any():
        return 0.0
    return float(df.loc[filtro, "Valor"].iloc[0])

def pct_linha(df: pd.DataFrame, linha_contains: str, mes: str) -> float:
    filtro = df["Linha DRE"].str.upper().str.contains(linha_contains.upper(), regex=False, na=False) & (df["Mês"] == mes)
    if not filtro.any():
        return 0.0
    return float(df.loc[filtro, "% Receita"].iloc[0])

def delta_mes(df: pd.DataFrame, linha_contains: str, meses: list[str]) -> tuple[float, str]:
    if len(meses) < 2:
        return 0.0, "Sem comparação"
    atual, anterior = meses[-1], meses[-2]
    v_atual = valor_linha(df, linha_contains, atual)
    v_ant = valor_linha(df, linha_contains, anterior)
    if v_ant == 0:
        return 0.0, f"{atual} vs {anterior}"
    return (v_atual - v_ant) / abs(v_ant), f"{atual} vs {anterior}"

def make_kpi(label: str, value: str, delta_value: float | None = None, delta_label: str = ""):
    if delta_value is None:
        delta_html = ""
    else:
        cls = "delta-pos" if delta_value >= 0 else "delta-neg"
        sinal = "+" if delta_value >= 0 else ""
        delta_html = f'<div class="delta {cls}">{sinal}{pct(delta_value)} • {delta_label}</div>'
    st.markdown(f"""
        <div class="kpi-card">
            <div class="label">{label}</div>
            <div class="value">{value}</div>
            {delta_html}
        </div>
        """, unsafe_allow_html=True)

def dre_pivot_html(dados: pd.DataFrame, meses: list[str], secao: str | None = None) -> str:
    base = dados.copy()
    if secao:
        base = base[base["Seção"].astype(str).str.upper() == secao.upper()]
    base = base[base["Mês"].isin(meses)]
    ordem = base[["Ordem", "Linha DRE", "Nível", "Tipo"]].drop_duplicates().sort_values("Ordem")
    header = "<tr><th>Linha DRE</th>" + "".join([f"<th>{m} Valor</th><th>{m} %</th>" for m in meses]) + "</tr>"
    linhas_html = []
    for _, row in ordem.iterrows():
        linha = row["Linha DRE"]
        tipo = row["Tipo"] if row["Tipo"] in ["subtotal_azul", "resultado_amarelo", "agrupador", "grupo_italico"] else "detalhe"
        nivel = int(row.get("Nível", 0))
        indent = "&nbsp;" * (nivel * 4)
        tds = [f"<td>{indent}{linha}</td>"]
        for mes in meses:
            rec = base[(base["Ordem"] == row["Ordem"]) & (base["Mês"] == mes)]
            val = float(rec["Valor"].iloc[0]) if not rec.empty else 0.0
            prc = float(rec["% Receita"].iloc[0]) if not rec.empty else 0.0
            if "MARGEM DE CONTRIBUIÇÃO PERCENTUAL" in str(linha).upper():
                tds.append(f"<td>{pct(val)}</td><td>{pct(prc)}</td>")
            else:
                tds.append(f"<td>{brl(val)}</td><td>{pct(prc)}</td>")
        linhas_html.append(f"<tr class='{tipo}'>" + "".join(tds) + "</tr>")
    return f"<div class='eirox-table-wrap'><table class='eirox-table'>{header}{''.join(linhas_html)}</table></div>"

def serie_linha(dados: pd.DataFrame, linha_contains: str, meses: list[str]) -> pd.DataFrame:
    return pd.DataFrame([{"Mês": m, "Valor": valor_linha(dados, linha_contains, m)} for m in meses])


# =========================================================
# MELHORIAS PREMIUM - PAINEL CEO / FARMACÊUTICO / PDF
# =========================================================
def valor_mes_linha(dados: pd.DataFrame, busca: str, mes: str) -> float:
    return valor_linha(dados, busca, mes)


def delta_percentual_linha(dados: pd.DataFrame, busca: str, meses: list[str]) -> float | None:
    if len(meses) < 2:
        return None
    atual = valor_mes_linha(dados, busca, meses[-1])
    anterior = valor_mes_linha(dados, busca, meses[-2])
    if anterior == 0:
        return None
    return (atual - anterior) / abs(anterior)


def seta_tendencia(delta: float | None) -> str:
    if delta is None:
        return "▬ 0,00%"
    if delta > 0:
        return f"▲ {pct(delta)}"
    if delta < 0:
        return f"▼ {pct(abs(delta))}"
    return "▬ 0,00%"


def classe_delta(delta: float | None) -> str:
    if delta is None or delta == 0:
        return "delta-neutral"
    return "delta-pos" if delta > 0 else "delta-neg"


def make_exec_card(titulo: str, valor: str, subtitulo: str = "", destaque: bool = False):
    borda = "rgba(255,210,31,.48)" if destaque else "rgba(0,168,255,.30)"
    fundo = "linear-gradient(180deg,#1e2630,#091728)" if destaque else "linear-gradient(180deg,#10223a,#091728)"
    st.markdown(
        f"""
        <div class="kpi-card" style="border-color:{borda}; background:{fundo};">
            <div class="label">{titulo}</div>
            <div class="value">{valor}</div>
            <div class="delta">{subtitulo}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def make_tendencia_card(nome: str, delta: float | None):
    cls = classe_delta(delta)
    st.markdown(
        f"""
        <div class="kpi-card" style="min-height:105px;">
            <div class="label">{nome}</div>
            <div class="value {cls}" style="font-size:1.75rem;">{seta_tendencia(delta)}</div>
            <div class="delta">comparativo mensal</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def semaforo(label: str, nivel: str, detalhe: str):
    cores = {
        "Saudável": ("🟢", "#32e875"),
        "Atenção": ("🟡", "#ffd21f"),
        "Crítico": ("🔴", "#ff5470"),
    }
    icone, cor = cores.get(nivel, ("⚪", "#9aa7b8"))
    st.markdown(
        f"""
        <div class="kpi-card" style="min-height:116px; border-color:{cor};">
            <div class="label">{label}</div>
            <div class="value" style="font-size:1.35rem;color:{cor};">{icone} {nivel}</div>
            <div class="delta">{detalhe}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def gerar_pdf_executivo(dados: pd.DataFrame, meses: list[str], logo_path: Path | None = None) -> bytes | None:
    """Gera PDF executivo mantendo o resumo do último mês e adicionando consolidado do período."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
    except Exception:
        return None

    def soma_linha(busca: str) -> float:
        return sum(valor_linha(dados, busca, m) for m in meses)

    def media_ratio_linha(busca: str) -> float:
        vals = [pct_linha(dados, busca, m) for m in meses]
        vals = [v for v in vals if v is not None]
        return sum(vals) / len(vals) if vals else 0.0

    def ratio_total(numerador: float, denominador: float) -> float:
        return numerador / denominador if denominador else 0.0

    def estilo_tabela(tbl: Table):
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0B2B45")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#F4F8FC")),
            ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#D5E2EF")),
            ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
            ("ALIGN", (1,1), (-1,-1), "RIGHT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EEF5FB")]),
        ]))
        return tbl

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.0*cm,
        leftMargin=1.0*cm,
        topMargin=0.8*cm,
        bottomMargin=0.8*cm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="EiroxSmall", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#4F5F70")))
    elementos = []

    if logo_path and logo_path.exists():
        try:
            img = Image(str(logo_path), width=4.2*cm, height=1.3*cm)
            elementos.append(img)
            elementos.append(Spacer(1, 0.2*cm))
        except Exception:
            pass

    ultimo = meses[-1]
    periodo_txt = f"{meses[0]} a {meses[-1]}"

    # =====================================================
    # Página 1 — resumo do último mês, preservando o modelo atual
    # =====================================================
    elementos.append(Paragraph("DRE Empresa Online - Relatório Executivo", styles["Title"]))
    elementos.append(Paragraph(f"Período: {periodo_txt} | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.35*cm))
    elementos.append(Paragraph(f"Resumo Executivo do Último Mês ({ultimo})", styles["Heading2"]))

    linhas = [
        ["Indicador", "Valor", "% Receita"],
        ["Receita Total", brl(valor_linha(dados, "RECEITA OPERACIONAL BRUTA", ultimo)), pct(pct_linha(dados, "RECEITA OPERACIONAL BRUTA", ultimo))],
        ["Lucro Bruto", brl(valor_linha(dados, "LUCRO BRUTO", ultimo)), pct(pct_linha(dados, "LUCRO BRUTO", ultimo))],
        ["EBITDA", brl(valor_linha(dados, "EBITDA", ultimo)), pct(pct_linha(dados, "EBITDA", ultimo))],
        ["Lucro Líquido", brl(valor_linha(dados, "LUCRO LÍQUIDO", ultimo)), pct(pct_linha(dados, "LUCRO LÍQUIDO", ultimo))],
        ["Posição Final Caixa", brl(valor_linha(dados, "POSIÇÃO FINAL", ultimo)), pct(pct_linha(dados, "POSIÇÃO FINAL", ultimo))],
    ]
    elementos.append(estilo_tabela(Table(linhas, colWidths=[11.5*cm, 6*cm, 4*cm])))
    elementos.append(Spacer(1, 0.5*cm))

    elementos.append(Paragraph("Resumo Farmacêutico do Último Mês", styles["Heading2"]))
    receita = valor_linha(dados, "RECEITA OPERACIONAL BRUTA", ultimo)
    estoque = valor_linha(dados, "ESTOQUE A CUSTO", ultimo)
    cmv = valor_linha(dados, "CUSTOS DAS VENDAS", ultimo)
    despesas_op = valor_linha(dados, "DESPESAS OPERACIONAIS", ultimo)
    pos_final = valor_linha(dados, "POSIÇÃO FINAL", ultimo)
    farmacia = [
        ["Indicador", "Resultado"],
        ["Estoque / Receita", f"{(estoque/receita):.2f}x".replace(".", ",") if receita else "0,00x"],
        ["Dias de Estoque", f"{(estoque/cmv*30):.1f} dias".replace(".", ",") if cmv else "0,0 dias"],
        ["Capital de Giro", brl(estoque - pos_final)],
        ["CMV %", pct(cmv/receita if receita else 0)],
        ["Despesa Operacional %", pct(despesas_op/receita if receita else 0)],
        ["EBITDA %", pct(pct_linha(dados, "EBITDA", ultimo))],
    ]
    elementos.append(estilo_tabela(Table(farmacia, colWidths=[11.5*cm, 10*cm])))

    # =====================================================
    # Página 2 — consolidado do período
    # =====================================================
    elementos.append(PageBreak())
    elementos.append(Paragraph("Resumo Consolidado do Período", styles["Title"]))
    elementos.append(Paragraph(f"Período consolidado: {periodo_txt}", styles["Normal"]))
    elementos.append(Spacer(1, 0.35*cm))

    receita_total = soma_linha("RECEITA OPERACIONAL BRUTA")
    lucro_bruto_total = soma_linha("LUCRO BRUTO")
    ebitda_total = soma_linha("EBITDA")
    lucro_liq_total = soma_linha("LUCRO LÍQUIDO")
    posicao_final_ultimo = valor_linha(dados, "POSIÇÃO FINAL", ultimo)
    estoque_medio = sum(valor_linha(dados, "ESTOQUE A CUSTO", m) for m in meses) / len(meses)
    posicao_media = sum(valor_linha(dados, "POSIÇÃO FINAL", m) for m in meses) / len(meses)
    capital_giro_medio = estoque_medio - posicao_media

    consolidado = [
        ["Indicador", "Consolidado", "% Receita"],
        ["Receita Acumulada", brl(receita_total), "100,00%"],
        ["Lucro Bruto Acumulado", brl(lucro_bruto_total), pct(ratio_total(lucro_bruto_total, receita_total))],
        ["EBITDA Acumulado", brl(ebitda_total), pct(ratio_total(ebitda_total, receita_total))],
        ["Lucro Líquido Acumulado", brl(lucro_liq_total), pct(ratio_total(lucro_liq_total, receita_total))],
        ["Posição Final Caixa", brl(posicao_final_ultimo), pct(ratio_total(posicao_final_ultimo, receita))],
        ["Margem EBITDA Consolidada", pct(ratio_total(ebitda_total, receita_total)), ""],
        ["Margem Líquida Consolidada", pct(ratio_total(lucro_liq_total, receita_total)), ""],
        ["Capital de Giro Médio", brl(capital_giro_medio), ""],
    ]
    elementos.append(estilo_tabela(Table(consolidado, colWidths=[11.5*cm, 6*cm, 4*cm])))
    elementos.append(Spacer(1, 0.55*cm))

    elementos.append(Paragraph("Evolução Mensal", styles["Heading2"]))
    evol_header = ["Indicador"] + meses
    evol_linhas = [
        ["Receita"] + [brl(valor_linha(dados, "RECEITA OPERACIONAL BRUTA", m)) for m in meses],
        ["Lucro Bruto"] + [brl(valor_linha(dados, "LUCRO BRUTO", m)) for m in meses],
        ["EBITDA"] + [brl(valor_linha(dados, "EBITDA", m)) for m in meses],
        ["Lucro Líquido"] + [brl(valor_linha(dados, "LUCRO LÍQUIDO", m)) for m in meses],
        ["Caixa Final"] + [brl(valor_linha(dados, "POSIÇÃO FINAL", m)) for m in meses],
    ]
    largura_primeira = 5.2*cm
    largura_mes = max(2.6*cm, (25.0*cm - largura_primeira) / max(len(meses), 1))
    elementos.append(estilo_tabela(Table([evol_header] + evol_linhas, colWidths=[largura_primeira] + [largura_mes]*len(meses))))

    # =====================================================
    # Página 3 — consolidado farmacêutico + parecer
    # =====================================================
    elementos.append(PageBreak())
    elementos.append(Paragraph("Consolidado Farmacêutico", styles["Title"]))
    elementos.append(Paragraph(f"Indicadores médios e acumulados do período: {periodo_txt}", styles["Normal"]))
    elementos.append(Spacer(1, 0.35*cm))

    cmv_total = soma_linha("CUSTOS DAS VENDAS")
    despesas_total = soma_linha("DESPESAS OPERACIONAIS")
    estoque_receita_medio = sum((valor_linha(dados, "ESTOQUE A CUSTO", m) / valor_linha(dados, "RECEITA OPERACIONAL BRUTA", m)) if valor_linha(dados, "RECEITA OPERACIONAL BRUTA", m) else 0 for m in meses) / len(meses)
    dias_estoque_medio = sum((valor_linha(dados, "ESTOQUE A CUSTO", m) / valor_linha(dados, "CUSTOS DAS VENDAS", m) * 30) if valor_linha(dados, "CUSTOS DAS VENDAS", m) else 0 for m in meses) / len(meses)

    farm_consolidado = [
        ["Indicador", "Consolidado / Média"],
        ["Estoque / Receita Médio", f"{estoque_receita_medio:.2f}x".replace(".", ",")],
        ["Dias de Estoque Médio", f"{dias_estoque_medio:.1f} dias".replace(".", ",")],
        ["Capital de Giro Médio", brl(capital_giro_medio)],
        ["CMV % Consolidado", pct(ratio_total(cmv_total, receita_total))],
        ["Despesa Operacional % Consolidada", pct(ratio_total(despesas_total, receita_total))],
        ["EBITDA % Consolidado", pct(ratio_total(ebitda_total, receita_total))],
    ]
    elementos.append(estilo_tabela(Table(farm_consolidado, colWidths=[11.5*cm, 10*cm])))
    elementos.append(Spacer(1, 0.55*cm))

    margem_ebitda = ratio_total(ebitda_total, receita_total)
    margem_liquida = ratio_total(lucro_liq_total, receita_total)
    if margem_ebitda >= 0.10 and margem_liquida >= 0.05:
        situacao = "saudável"
        comentario = "A empresa apresenta geração operacional positiva e margem consistente para o período analisado."
    elif margem_ebitda >= 0.05:
        situacao = "em atenção"
        comentario = "A operação apresenta resultado positivo, porém com necessidade de acompanhamento de margem, estoque e despesas."
    else:
        situacao = "crítica"
        comentario = "A operação requer atenção imediata sobre margem, estrutura de despesas e capital imobilizado."

    elementos.append(Paragraph("Parecer Executivo Automático", styles["Heading2"]))
    parecer = (
        f"Receita acumulada do período: <b>{brl(receita_total)}</b>.<br/>"
        f"EBITDA acumulado: <b>{brl(ebitda_total)}</b>, com margem consolidada de <b>{pct(margem_ebitda)}</b>.<br/>"
        f"Lucro líquido acumulado: <b>{brl(lucro_liq_total)}</b>, com margem líquida de <b>{pct(margem_liquida)}</b>.<br/>"
        f"O estoque médio representa <b>{estoque_receita_medio:.2f}x</b> a receita mensal média e equivale a aproximadamente <b>{dias_estoque_medio:.1f} dias</b> de CMV.<br/>"
        f"Situação financeira do período: <b>{situacao.upper()}</b>. {comentario}"
    ).replace(".", ",", 1) if False else (
        f"Receita acumulada do período: <b>{brl(receita_total)}</b>.<br/>"
        f"EBITDA acumulado: <b>{brl(ebitda_total)}</b>, com margem consolidada de <b>{pct(margem_ebitda)}</b>.<br/>"
        f"Lucro líquido acumulado: <b>{brl(lucro_liq_total)}</b>, com margem líquida de <b>{pct(margem_liquida)}</b>.<br/>"
        f"O estoque médio representa <b>{estoque_receita_medio:.2f}x</b> a receita mensal média e equivale a aproximadamente <b>{dias_estoque_medio:.1f} dias</b> de CMV.<br/>"
        f"Situação financeira do período: <b>{situacao.upper()}</b>. {comentario}"
    )
    parecer = parecer.replace(".", ",") if False else parecer
    elementos.append(Paragraph(parecer, styles["Normal"]))
    elementos.append(Spacer(1, 0.35*cm))
    elementos.append(Paragraph("EIROX FINANCIAL ANALYTICS • Relatório Executivo gerado automaticamente", styles["EiroxSmall"]))

    doc.build(elementos)
    return buffer.getvalue()


def gerar_excel_executivo(dados: pd.DataFrame, meses: list[str], checks: pd.DataFrame | None = None, nao_classificados: pd.DataFrame | None = None, logo_path: Path | None = None) -> bytes:
    """Gera Excel executivo visual, espelhando o dashboard Eirox: tema dark, logo, cards, DRE com destaques e abas premium."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.drawing.image import Image as XLImage

    buffer = io.BytesIO()
    ultimo = meses[-1]

    # Paleta Eirox / Dashboard
    C_BG = "07111F"
    C_PANEL = "0B1728"
    C_PANEL2 = "10223A"
    C_HEADER = "0B2B45"
    C_LINE = "20334D"
    C_BLUE = "CFE3F8"
    C_YELLOW = "FFD21F"
    C_GROUP = "1B2534"
    C_WHITE = "FFFFFF"
    C_MUTED = "9FB5CC"
    C_GREEN = "32E875"
    C_RED = "FF5470"
    C_BLACK = "00101E"

    money_fmt = '"R$" #,##0.00'
    pct_fmt = '0.00%'
    num_fmt = '#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "Painel CEO"
    for name in ["DRE Gerencial", "Evolução Mensal", "Indicadores Farma", "Auditoria"]:
        wb.create_sheet(name)

    thin = Side(style="thin", color=C_LINE)
    light = Side(style="thin", color="D5E2EF")

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def set_dark_sheet(ws):
        ws.sheet_view.showGridLines = False
        for r in range(1, 90):
            for c in range(1, 22):
                ws.cell(r, c).fill = fill(C_BG)
                ws.cell(r, c).font = Font(color=C_WHITE, name="Calibri")

    def add_logo(ws, cell="A1", width=135):
        if logo_path and Path(logo_path).exists():
            try:
                img = XLImage(str(logo_path))
                img.width = width
                img.height = int(width * 0.32)
                ws.add_image(img, cell)
            except Exception:
                pass

    def title(ws, text, subtitle=None):
        ws.merge_cells("A1:H2")
        ws["A1"] = text
        ws["A1"].font = Font(color=C_WHITE, bold=True, size=24)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A3:H3")
        ws["A3"] = subtitle or "EIROX FINANCIAL ANALYTICS"
        ws["A3"].font = Font(color="6CC7FF", bold=True, size=11)
        ws["A3"].alignment = Alignment(horizontal="center")
        for row in range(1, 4):
            for col in range(1, 9):
                ws.cell(row, col).fill = fill(C_PANEL)
                ws.cell(row, col).border = Border(bottom=thin)

    def style_cell(cell, bg=C_PANEL, font=C_WHITE, bold=False, size=11, align="left", numfmt=None):
        cell.fill = fill(bg)
        cell.font = Font(color=font, bold=bold, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if numfmt:
            cell.number_format = numfmt

    def card(ws, row, col, label, value, pct_value=None, color=C_PANEL2):
        ws.merge_cells(start_row=row, start_column=col, end_row=row+2, end_column=col+1)
        c = ws.cell(row, col)
        c.value = f"{label}\n{value}" + (f"\n{pct_value}" if pct_value else "")
        c.fill = fill(color)
        c.font = Font(color=C_WHITE, bold=True, size=13)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def valor(busca, mes=ultimo):
        return valor_linha(dados, busca, mes)

    def porcent(busca, mes=ultimo):
        return pct_linha(dados, busca, mes)

    def soma(busca):
        return sum(valor_linha(dados, busca, m) for m in meses)

    def ratio(a, b):
        return a / b if b else 0

    # =====================================================
    # Aba Painel CEO
    # =====================================================
    ws = wb["Painel CEO"]
    set_dark_sheet(ws)
    add_logo(ws, "A1", 120)
    title(ws, "DRE Empresa Online", f"Painel CEO Premium • Período: {meses[0]} a {meses[-1]} • Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    receita = valor("RECEITA OPERACIONAL BRUTA")
    lucro_bruto = valor("LUCRO BRUTO")
    ebitda = valor("EBITDA")
    lucro_liq = valor("LUCRO LÍQUIDO")
    caixa = valor("POSIÇÃO FINAL")
    margem_ebitda = ratio(ebitda, receita)
    margem_liq = ratio(lucro_liq, receita)

    cards = [
        ("Receita Total", brl(receita), "100,00%"),
        ("Lucro Bruto", brl(lucro_bruto), pct(ratio(lucro_bruto, receita))),
        ("EBITDA", brl(ebitda), pct(margem_ebitda)),
        ("Lucro Líquido", brl(lucro_liq), pct(margem_liq)),
        ("Posição Final Caixa", brl(caixa), pct(ratio(caixa, receita))),
        ("Margem EBITDA", pct(margem_ebitda), None),
        ("Margem Líquida", pct(margem_liq), None),
    ]
    pos = [(5,1),(5,3),(5,5),(5,7),(9,1),(9,3),(9,5)]
    for (r,c), (lab,val,pc) in zip(pos, cards):
        card(ws, r, c, lab, val, pc)

    ws["A14"] = "Comparativo Mensal"
    style_cell(ws["A14"], C_HEADER, C_WHITE, True, 14)
    ws.merge_cells("A14:H14")
    headers = ["Indicador", "Mês Anterior", "Último Mês", "Variação", "Tendência"]
    for i, h in enumerate(headers, 1):
        style_cell(ws.cell(15, i), C_HEADER, C_WHITE, True, 11, "center")
    comparativos = [
        ("Receita", "RECEITA OPERACIONAL BRUTA"),
        ("Lucro Bruto", "LUCRO BRUTO"),
        ("EBITDA", "EBITDA"),
        ("Lucro Líquido", "LUCRO LÍQUIDO"),
        ("Caixa Final", "POSIÇÃO FINAL"),
    ]
    ant = meses[-2] if len(meses) > 1 else meses[-1]
    for idx, (lab, busca) in enumerate(comparativos, 16):
        v_ant = valor_linha(dados, busca, ant)
        v_atual = valor_linha(dados, busca, ultimo)
        delta = (v_atual - v_ant) / abs(v_ant) if v_ant else 0
        tendencia = "▲" if delta >= 0 else "▼"
        vals = [lab, v_ant, v_atual, delta, tendencia]
        for col, x in enumerate(vals, 1):
            cell = ws.cell(idx, col, x)
            style_cell(cell, C_PANEL, C_WHITE, col == 1, 11, "right" if col in [2,3,4] else "center")
            if col in [2,3]: cell.number_format = money_fmt
            if col == 4: cell.number_format = pct_fmt
            if col == 5:
                cell.font = Font(color=C_GREEN if delta >= 0 else C_RED, bold=True, size=14)

    ws["A23"] = "Semáforo Financeiro"
    style_cell(ws["A23"], C_HEADER, C_WHITE, True, 14)
    ws.merge_cells("A23:H23")
    semaforos = [
        ("Margem", margem_ebitda, "Saudável" if margem_ebitda >= .10 else "Atenção" if margem_ebitda >= .05 else "Crítico"),
        ("Caixa", caixa, "Saudável" if caixa > 0 else "Crítico"),
        ("Estoque", ratio(valor("ESTOQUE A CUSTO"), receita), "Saudável" if ratio(valor("ESTOQUE A CUSTO"), receita) <= 2.2 else "Atenção"),
        ("Despesas", porcent("DESPESAS OPERACIONAIS"), "Saudável" if porcent("DESPESAS OPERACIONAIS") <= .28 else "Atenção" if porcent("DESPESAS OPERACIONAIS") <= .35 else "Crítico"),
    ]
    for j, (lab, valx, status) in enumerate(semaforos, 1):
        col = 1 + (j-1)*2
        cor = C_GREEN if status == "Saudável" else C_YELLOW if status == "Atenção" else C_RED
        card(ws, 25, col, lab, status, None, color=C_PANEL2)
        ws.cell(25, col).font = Font(color=cor, bold=True, size=13)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A15"

    # =====================================================
    # Aba DRE Gerencial — igual ao dashboard
    # =====================================================
    ws = wb["DRE Gerencial"]
    set_dark_sheet(ws)
    add_logo(ws, "A1", 110)
    title(ws, "DRE Gerencial", "Modelo completo: DRE + Conciliação de Fluxo de Caixa + Indicadores + Ponto de Equilíbrio")

    start_row = 5
    headers = ["Linha DRE"] + [x for m in meses for x in [f"{m} Valor", f"{m} %"]]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        style_cell(cell, "141B27", "C9D3DF", True, 11, "right" if c > 1 else "left")

    ordem = dados[["Ordem", "Seção", "Linha DRE", "Nível", "Tipo"]].drop_duplicates().sort_values("Ordem")
    r = start_row + 1
    for _, row in ordem.iterrows():
        linha = row["Linha DRE"]
        tipo = row["Tipo"] if row["Tipo"] in ["subtotal_azul", "resultado_amarelo", "agrupador", "grupo_italico"] else "detalhe"
        nivel = int(row.get("Nível", 0))
        if tipo == "resultado_amarelo": bg, fc, bold = C_YELLOW, C_BLACK, True
        elif tipo == "subtotal_azul": bg, fc, bold = C_BLUE, C_BLACK, True
        elif tipo in ["agrupador", "grupo_italico"]: bg, fc, bold = C_GROUP, C_WHITE, True
        else: bg, fc, bold = C_PANEL, C_WHITE, False
        c0 = ws.cell(r, 1, linha)
        style_cell(c0, bg, fc, bold, 11, "left")
        c0.alignment = Alignment(horizontal="left", vertical="center", indent=nivel, wrap_text=True)
        col = 2
        for m in meses:
            rec = dados[(dados["Ordem"] == row["Ordem"]) & (dados["Mês"] == m)]
            val = float(rec["Valor"].iloc[0]) if not rec.empty else 0.0
            prc = float(rec["% Receita"].iloc[0]) if not rec.empty else 0.0
            cell_v = ws.cell(r, col, val)
            cell_p = ws.cell(r, col+1, prc)
            style_cell(cell_v, bg, fc, bold, 11, "right", money_fmt)
            style_cell(cell_p, bg, fc, bold, 11, "right", pct_fmt)
            if "MARGEM DE CONTRIBUIÇÃO PERCENTUAL" in str(linha).upper():
                cell_v.number_format = pct_fmt
            col += 2
        ws.row_dimensions[r].height = 28 if tipo in ["subtotal_azul", "resultado_amarelo"] else 23
        r += 1
    ws.freeze_panes = "B6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{r-1}"
    ws.column_dimensions["A"].width = 56
    for c in range(2, len(headers)+1):
        ws.column_dimensions[get_column_letter(c)].width = 17 if c % 2 == 0 else 11

    # =====================================================
    # Aba Evolução Mensal com gráfico
    # =====================================================
    ws = wb["Evolução Mensal"]
    set_dark_sheet(ws)
    add_logo(ws, "A1", 105)
    title(ws, "Evolução Mensal", f"Indicadores de {meses[0]} a {meses[-1]}")
    indicadores = [
        ("Receita Bruta", "RECEITA OPERACIONAL BRUTA"),
        ("Lucro Bruto", "LUCRO BRUTO"),
        ("EBITDA", "EBITDA"),
        ("Lucro Líquido", "LUCRO LÍQUIDO"),
        ("Posição Final", "POSIÇÃO FINAL"),
        ("Estoque a Custo", "ESTOQUE A CUSTO"),
    ]
    header = ["Indicador"] + meses + ["Total/Consolidado"]
    for c, h in enumerate(header, 1):
        style_cell(ws.cell(5, c, h), C_HEADER, C_WHITE, True, 11, "center")
    for r_idx, (lab, busca) in enumerate(indicadores, 6):
        style_cell(ws.cell(r_idx, 1, lab), C_PANEL, C_WHITE, True)
        total = 0
        for i, m in enumerate(meses, 2):
            valm = valor_linha(dados, busca, m)
            total += valm
            style_cell(ws.cell(r_idx, i, valm), C_PANEL, C_WHITE, False, 11, "right", money_fmt)
        style_cell(ws.cell(r_idx, len(meses)+2, total), C_PANEL2, C_WHITE, True, 11, "right", money_fmt)
    ws.freeze_panes = "B6"
    ws.column_dimensions["A"].width = 24
    for c in range(2, len(header)+1): ws.column_dimensions[get_column_letter(c)].width = 16
    try:
        chart = LineChart()
        chart.title = "Evolução dos Indicadores"
        chart.y_axis.title = "Valor"
        chart.x_axis.title = "Mês"
        data = Reference(ws, min_col=2, max_col=1+len(meses), min_row=5, max_row=10)
        cats = Reference(ws, min_col=2, max_col=1+len(meses), min_row=5, max_row=5)
        chart.add_data(data, titles_from_data=False, from_rows=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 24
        ws.add_chart(chart, "A15")
    except Exception:
        pass

    # =====================================================
    # Aba Indicadores Farma
    # =====================================================
    ws = wb["Indicadores Farma"]
    set_dark_sheet(ws)
    add_logo(ws, "A1", 105)
    title(ws, "Dashboard Farmacêutico", "Indicadores específicos para rede de farmácias")
    indicadores_farma = [
        ("Estoque / Receita", lambda m: ratio(valor_linha(dados, "ESTOQUE A CUSTO", m), valor_linha(dados, "RECEITA OPERACIONAL BRUTA", m)), "0.00x"),
        ("Dias de Estoque", lambda m: ratio(valor_linha(dados, "ESTOQUE A CUSTO", m), valor_linha(dados, "CUSTOS DAS VENDAS", m)) * 30 if valor_linha(dados, "CUSTOS DAS VENDAS", m) else 0, '0.0 "dias"'),
        ("Capital de Giro", lambda m: valor_linha(dados, "ESTOQUE A CUSTO", m) - valor_linha(dados, "POSIÇÃO FINAL", m), money_fmt),
        ("CMV %", lambda m: ratio(valor_linha(dados, "CUSTOS DAS VENDAS", m), valor_linha(dados, "RECEITA OPERACIONAL BRUTA", m)), pct_fmt),
        ("Despesa Operacional %", lambda m: ratio(valor_linha(dados, "DESPESAS OPERACIONAIS", m), valor_linha(dados, "RECEITA OPERACIONAL BRUTA", m)), pct_fmt),
        ("EBITDA %", lambda m: ratio(valor_linha(dados, "EBITDA", m), valor_linha(dados, "RECEITA OPERACIONAL BRUTA", m)), pct_fmt),
    ]
    header = ["Indicador"] + meses + ["Média/Consolidado"]
    for c, h in enumerate(header, 1):
        style_cell(ws.cell(5, c, h), C_HEADER, C_WHITE, True, 11, "center")
    for r_idx, (lab, fn, fmt) in enumerate(indicadores_farma, 6):
        vals = [fn(m) for m in meses]
        media = sum(vals) / len(vals) if vals else 0
        style_cell(ws.cell(r_idx, 1, lab), C_PANEL, C_WHITE, True)
        for i, v in enumerate(vals, 2):
            style_cell(ws.cell(r_idx, i, v), C_PANEL, C_WHITE, False, 11, "right", fmt)
        style_cell(ws.cell(r_idx, len(meses)+2, media), C_PANEL2, C_WHITE, True, 11, "right", fmt)
    ws.freeze_panes = "B6"
    ws.column_dimensions["A"].width = 26
    for c in range(2, len(header)+1): ws.column_dimensions[get_column_letter(c)].width = 16

    # =====================================================
    # Aba Auditoria
    # =====================================================
    ws = wb["Auditoria"]
    set_dark_sheet(ws)
    add_logo(ws, "A1", 105)
    title(ws, "Auditoria DRE", "Checagem das bases e itens não classificados")
    row = 5
    for c, h in enumerate(["Checagem", "Valor"], 1):
        style_cell(ws.cell(row, c, h), C_HEADER, C_WHITE, True, 11, "center")
    row += 1
    if checks is not None and not checks.empty:
        for _, rr in checks.iterrows():
            style_cell(ws.cell(row, 1, rr.get("Checagem", "")), C_PANEL, C_WHITE, True)
            style_cell(ws.cell(row, 2, rr.get("Valor", "")), C_PANEL, C_WHITE, False)
            row += 1
    else:
        style_cell(ws.cell(row, 1, "Auditoria"), C_PANEL, C_WHITE, True)
        style_cell(ws.cell(row, 2, "Sem auditoria disponível"), C_PANEL, C_WHITE, False)
        row += 1
    row += 2
    ws.cell(row, 1, "Não Classificados")
    style_cell(ws.cell(row, 1), C_HEADER, C_WHITE, True, 13)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    if nao_classificados is not None and not nao_classificados.empty:
        cols = list(nao_classificados.columns)[:8]
        for c, h in enumerate(cols, 1):
            style_cell(ws.cell(row, c, h), C_HEADER, C_WHITE, True, 11, "center")
        row += 1
        for _, rr in nao_classificados.head(1000).iterrows():
            for c, h in enumerate(cols, 1):
                style_cell(ws.cell(row, c, rr.get(h, "")), C_PANEL, C_WHITE, False)
            row += 1
    else:
        style_cell(ws.cell(row, 1, "Nenhum item não classificado encontrado."), C_PANEL, C_GREEN, True)
    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 24

    # Ajustes finais de impressão / página
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# =========================================================
# LOGIN
# =========================================================
def criar_token_login(usuario: str) -> str:
    usuario = str(usuario).strip().lower()
    expira_em = int(time.time()) + LOGIN_DURACAO_SEGUNDOS
    payload = f"{usuario}|{expira_em}"
    assinatura = hmac.new(LOGIN_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{payload}|{assinatura}"


def validar_token_login(token: str) -> str | None:
    try:
        usuario, expira_em, assinatura = str(token).split("|", 2)
        payload = f"{usuario}|{expira_em}"
        assinatura_correta = hmac.new(LOGIN_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(assinatura, assinatura_correta):
            return None
        if int(expira_em) < int(time.time()):
            return None
        if usuario not in USUARIOS:
            return None
        return usuario
    except Exception:
        return None


def autenticar_usuario(usuario: str):
    user = USUARIOS[usuario]
    st.session_state["autenticado"] = True
    st.session_state["usuario"] = usuario
    st.session_state["nome_usuario"] = user["nome"]
    st.session_state["perfil"] = user["perfil"]


def tela_login() -> bool:
    if st.session_state.get("autenticado", False):
        return True

    # Se o usuário já entrou neste navegador, mantém o acesso durante o período definido.
    token_salvo = st.query_params.get("auth", "")
    usuario_token = validar_token_login(token_salvo) if token_salvo else None
    if usuario_token:
        autenticar_usuario(usuario_token)
        return True

    logo_path_login = encontrar_arquivo(POSSIVEIS_LOGOS)
    logo_html = ""
    if logo_path_login:
        b64 = img_base64(logo_path_login)
        logo_html = f"<div style='text-align:center;'><img src='data:image/png;base64,{b64}' width='190'></div>"
    st.markdown(f"""
        <div class='login-card'>
            {logo_html}
            <div class='login-title'>Eirox DRE Online</div>
            <div class='login-sub'>Acesso ao Painel Financeiro Enterprise</div>
        </div>
        """, unsafe_allow_html=True)
    with st.form("form_login", clear_on_submit=False):
        usuario = st.text_input("Usuário")
        senha = st.text_input("Senha", type="password")
        lembrar = st.checkbox("Manter conectado neste navegador", value=True)
        entrar = st.form_submit_button("Entrar")
    if entrar:
        usuario_limpo = str(usuario).strip().lower()
        user = USUARIOS.get(usuario_limpo)
        if user and senha == user["senha"]:
            autenticar_usuario(usuario_limpo)
            if lembrar:
                st.query_params["auth"] = criar_token_login(usuario_limpo)
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")
    st.stop()

def botao_logout_sidebar():
    nome = st.session_state.get("nome_usuario", "Usuário")
    perfil = st.session_state.get("perfil", "")
    st.sidebar.markdown("<div class='sidebar-section'>Acesso</div>", unsafe_allow_html=True)
    st.sidebar.caption(f"Usuário: {nome}")
    st.sidebar.caption(f"Perfil: {perfil}")
    if st.sidebar.button("Sair", use_container_width=True):
        for k in ["autenticado", "usuario", "nome_usuario", "perfil"]:
            st.session_state.pop(k, None)
        st.query_params.clear()
        st.rerun()

tela_login()

# =========================================================
# CARREGAMENTO
# =========================================================
logo_path = encontrar_arquivo(POSSIVEIS_LOGOS)

if logo_path:
    b64 = img_base64(logo_path)
    st.sidebar.markdown(f"<div class='sidebar-logo'><img src='data:image/png;base64,{b64}' width='210'></div>", unsafe_allow_html=True)
else:
    st.sidebar.markdown("<div class='sidebar-title'>EIROX</div>", unsafe_allow_html=True)

st.sidebar.markdown("<div class='sidebar-title'>DRE Financeiro Online</div>", unsafe_allow_html=True)
st.sidebar.markdown("<div class='sidebar-subtitle'>Enterprise Premium</div>", unsafe_allow_html=True)
botao_logout_sidebar()

st.sidebar.markdown("<div class='sidebar-section'>Navegação</div>", unsafe_allow_html=True)
pagina = st.sidebar.radio(
    "",
    ["📊 Painel CEO", "📑 DRE Gerencial", "📈 Evolução Mensal", "🏪 Resultado por Loja", "⚖️ Resultados Estratégicos", "⚠️ Auditoria DRE"],
    label_visibility="collapsed",
)

try:
    dados, resumo_loja, nao_classificados, checks, status_pastas = carregar_base_dre()
except Exception as e:
    st.markdown("<div class='hero'><h1>DRE Empresa Online</h1><p>Erro ao carregar base.</p></div>", unsafe_allow_html=True)
    st.error(f"Erro ao ler a base: {e}")
    st.stop()

all_months = sorted([m for m in dados["Mês"].dropna().unique().tolist() if str(m).lower() not in ["sem mês", "sem mes", "nan", "none", ""]], key=mes_sort_key)
default_months = meses_fechados_ano_atual(all_months) or all_months
st.sidebar.markdown("<div class='sidebar-section'>Filtros</div>", unsafe_allow_html=True)
meses_sel = st.sidebar.multiselect("Meses", all_months, default=default_months)
meses_sel = sorted(meses_sel, key=mes_sort_key)
if not meses_sel:
    st.warning("Selecione pelo menos um mês.")
    st.stop()

modo = "DINÂMICO"
if not checks.empty and "Modo de leitura" in checks["Checagem"].astype(str).values:
    modo = str(checks.loc[checks["Checagem"].astype(str) == "Modo de leitura", "Valor"].iloc[0])

st.sidebar.markdown("<div class='sidebar-section'>Status</div>", unsafe_allow_html=True)
st.sidebar.caption(f"Modo: {modo}")
st.sidebar.caption(f"Atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
st.sidebar.markdown("<div class='footer'>EIROX FINANCIAL ANALYTICS<br>Versão Enterprise Premium</div>", unsafe_allow_html=True)

ultimo_mes = meses_sel[-1]

logo_top = ""
if logo_path:
    b64 = img_base64(logo_path)
    logo_top = f"<img src='data:image/png;base64,{b64}' width='118' style='margin-bottom:12px;'>"

st.markdown(f"""
    <div class="hero" style="text-align:center;">
        {logo_top}
        <h1>DRE Empresa Online</h1>
        <p>Dashboard financeiro gerencial • DRE recalculado pelas pastas • Indicadores executivos premium</p>
        <div class="small-meta">Período filtrado: {meses_sel[0]} a {meses_sel[-1]} • Atualizado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
    </div>
    """, unsafe_allow_html=True)

# KPIs fixos
cols = st.columns(5)
with cols[0]:
    d, dl = delta_mes(dados, "RECEITA OPERACIONAL BRUTA", meses_sel)
    make_kpi("Receita Bruta", brl(valor_linha(dados, "RECEITA OPERACIONAL BRUTA", ultimo_mes)), d if len(meses_sel) > 1 else None, dl)
with cols[1]:
    d, dl = delta_mes(dados, "RECEITA OPERACIONAL LÍQUIDA", meses_sel)
    make_kpi("Receita Líquida", brl(valor_linha(dados, "RECEITA OPERACIONAL LÍQUIDA", ultimo_mes)), d if len(meses_sel) > 1 else None, dl)
with cols[2]:
    d, dl = delta_mes(dados, "LUCRO BRUTO", meses_sel)
    make_kpi("Lucro Bruto", brl(valor_linha(dados, "LUCRO BRUTO", ultimo_mes)), d if len(meses_sel) > 1 else None, dl)
with cols[3]:
    d, dl = delta_mes(dados, "EBITDA", meses_sel)
    make_kpi("EBITDA", brl(valor_linha(dados, "EBITDA", ultimo_mes)), d if len(meses_sel) > 1 else None, dl)
with cols[4]:
    d, dl = delta_mes(dados, "LUCRO LÍQUIDO", meses_sel)
    make_kpi("Lucro Líquido", brl(valor_linha(dados, "LUCRO LÍQUIDO", ultimo_mes)), d if len(meses_sel) > 1 else None, dl)

st.markdown("<br>", unsafe_allow_html=True)

# =========================================================
# PÁGINAS
# =========================================================
if pagina == "📊 Painel CEO":
    st.markdown("<div class='section-title'>📊 Painel CEO Premium</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-caption'>Visão executiva com cards, tendências mensais, semáforo financeiro, indicadores farmacêuticos e exportação PDF.</div>", unsafe_allow_html=True)

    receita_total = valor_linha(dados, "RECEITA OPERACIONAL BRUTA", ultimo_mes)
    lucro_bruto = valor_linha(dados, "LUCRO BRUTO", ultimo_mes)
    ebitda = valor_linha(dados, "EBITDA", ultimo_mes)
    lucro_liquido = valor_linha(dados, "LUCRO LÍQUIDO", ultimo_mes)
    posicao_final = valor_linha(dados, "POSIÇÃO FINAL", ultimo_mes)
    margem_ebitda = pct_linha(dados, "EBITDA", ultimo_mes)
    margem_liquida = pct_linha(dados, "LUCRO LÍQUIDO", ultimo_mes)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        make_exec_card("Receita Total", brl(receita_total), f"{ultimo_mes}")
    with c2:
        make_exec_card("Lucro Bruto", brl(lucro_bruto), pct(pct_linha(dados, "LUCRO BRUTO", ultimo_mes)))
    with c3:
        make_exec_card("EBITDA", brl(ebitda), pct(margem_ebitda), destaque=True)
    with c4:
        make_exec_card("Lucro Líquido", brl(lucro_liquido), pct(margem_liquida), destaque=True)

    c5, c6, c7 = st.columns(3)
    with c5:
        make_exec_card("Posição Final Caixa", brl(posicao_final), pct(pct_linha(dados, "POSIÇÃO FINAL", ultimo_mes)), destaque=True)
    with c6:
        make_exec_card("Margem EBITDA", pct(margem_ebitda), "EBITDA / Receita")
    with c7:
        make_exec_card("Margem Líquida", pct(margem_liquida), "Lucro Líquido / Receita")

    st.markdown("<div class='section-title'>📈 Comparativo Mensal</div>", unsafe_allow_html=True)
    t1, t2, t3 = st.columns(3)
    with t1:
        make_tendencia_card("Receita", delta_percentual_linha(dados, "RECEITA OPERACIONAL BRUTA", meses_sel))
    with t2:
        make_tendencia_card("Lucro", delta_percentual_linha(dados, "LUCRO LÍQUIDO", meses_sel))
    with t3:
        make_tendencia_card("EBITDA", delta_percentual_linha(dados, "EBITDA", meses_sel))

    st.markdown("<div class='section-title'>🚦 Semáforo Financeiro</div>", unsafe_allow_html=True)
    receita = receita_total or 0
    estoque = valor_linha(dados, "ESTOQUE A CUSTO", ultimo_mes)
    despesas_op = valor_linha(dados, "DESPESAS OPERACIONAIS", ultimo_mes)
    margem = margem_ebitda
    caixa_pct = (posicao_final / receita) if receita else 0
    estoque_receita = (estoque / receita) if receita else 0
    despesas_pct = (despesas_op / receita) if receita else 0

    status_margem = "Saudável" if margem >= 0.10 else ("Atenção" if margem >= 0.05 else "Crítico")
    status_caixa = "Saudável" if caixa_pct >= 0.05 else ("Atenção" if caixa_pct >= 0 else "Crítico")
    status_estoque = "Saudável" if estoque_receita <= 2.0 else ("Atenção" if estoque_receita <= 3.0 else "Crítico")
    status_despesas = "Saudável" if despesas_pct <= 0.25 else ("Atenção" if despesas_pct <= 0.35 else "Crítico")

    s1, s2, s3, s4 = st.columns(4)
    with s1:
        semaforo("Margem", status_margem, f"EBITDA: {pct(margem)}")
    with s2:
        semaforo("Caixa", status_caixa, f"Posição Final: {pct(caixa_pct)}")
    with s3:
        semaforo("Estoque", status_estoque, f"Estoque/Receita: {estoque_receita:.2f}x".replace(".", ","))
    with s4:
        semaforo("Despesas", status_despesas, f"Despesas Op.: {pct(despesas_pct)}")

    st.markdown("<div class='section-title'>💊 Dashboard Farmacêutico</div>", unsafe_allow_html=True)
    cmv = valor_linha(dados, "CUSTOS DAS VENDAS", ultimo_mes)
    dias_estoque = (estoque / cmv * 30) if cmv else 0
    capital_giro = estoque - posicao_final
    f1, f2, f3 = st.columns(3)
    with f1:
        make_exec_card("Estoque / Receita", f"{estoque_receita:.2f}x".replace(".", ","), "capital imobilizado em estoque")
    with f2:
        make_exec_card("Dias de Estoque", f"{dias_estoque:.1f} dias".replace(".", ","), "estoque / CMV x 30")
    with f3:
        make_exec_card("Capital de Giro", brl(capital_giro), "estoque menos posição final")
    f4, f5, f6 = st.columns(3)
    with f4:
        make_exec_card("CMV %", pct(cmv/receita if receita else 0), "custo sobre receita")
    with f5:
        make_exec_card("Despesa Operacional %", pct(despesas_pct), "despesas sobre receita")
    with f6:
        make_exec_card("EBITDA %", pct(margem), "resultado operacional")

    st.markdown("<div class='section-title'>📊 Evolução dos principais resultados</div>", unsafe_allow_html=True)
    evol = []
    for nome, busca in [("Receita Bruta", "RECEITA OPERACIONAL BRUTA"), ("Lucro Bruto", "LUCRO BRUTO"), ("EBITDA", "EBITDA"), ("Lucro Líquido", "LUCRO LÍQUIDO"), ("Posição Final", "POSIÇÃO FINAL")]:
        tmp = serie_linha(dados, busca, meses_sel)
        tmp["Indicador"] = nome
        evol.append(tmp)
    evol_df = pd.concat(evol, ignore_index=True)
    fig = px.line(evol_df, x="Mês", y="Valor", color="Indicador", markers=True, template="plotly_dark")
    fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", height=430, margin=dict(l=20, r=20, t=30, b=20))
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("<div class='section-title'>📥 Exportação Executiva</div>", unsafe_allow_html=True)
    col_pdf, col_excel = st.columns(2)
    with col_pdf:
        pdf_bytes = gerar_pdf_executivo(dados, meses_sel, logo_path)
        if pdf_bytes:
            st.download_button(
                "📄 Exportar PDF Executivo",
                data=pdf_bytes,
                file_name=f"Relatorio_Executivo_DRE_Eirox_{ultimo_mes.replace('/', '_')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        else:
            st.warning("Para exportar em PDF no ambiente online, adicione reportlab ao requirements.txt.")
    with col_excel:
        excel_bytes = gerar_excel_executivo(dados, meses_sel, checks, nao_classificados, logo_path)
        st.download_button(
            "📊 Exportar Excel Executivo",
            data=excel_bytes,
            file_name=f"Relatorio_Executivo_DRE_Eirox_{ultimo_mes.replace('/', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

elif pagina == "📑 DRE Gerencial":
    st.markdown("<div class='section-title'>📑 DRE Gerencial</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-caption'>Modelo completo: DRE + Conciliação de Fluxo de Caixa + Indicadores + Ponto de Equilíbrio.</div>", unsafe_allow_html=True)
    st.markdown(dre_pivot_html(dados, meses_sel, secao=None), unsafe_allow_html=True)

elif pagina == "📈 Evolução Mensal":
    st.markdown("<div class='section-title'>📈 Evolução Mensal</div>", unsafe_allow_html=True)
    indicadores = {
        "Receita Bruta": "RECEITA OPERACIONAL BRUTA",
        "Receita Líquida": "RECEITA OPERACIONAL LÍQUIDA",
        "Lucro Bruto": "LUCRO BRUTO",
        "EBITDA": "EBITDA",
        "Lucro Líquido": "LUCRO LÍQUIDO",
        "CMV": "CUSTOS DAS VENDAS",
        "Estoque a Custo": "ESTOQUE A CUSTO",
    }
    escolha = st.multiselect("Indicadores", list(indicadores.keys()), default=["Receita Bruta", "EBITDA", "Lucro Líquido"])
    evol = []
    for nome in escolha:
        tmp = serie_linha(dados, indicadores[nome], meses_sel)
        tmp["Indicador"] = nome
        evol.append(tmp)
    if evol:
        evol_df = pd.concat(evol, ignore_index=True)
        fig = px.bar(evol_df, x="Mês", y="Valor", color="Indicador", barmode="group", template="plotly_dark")
        fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", height=480)
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Selecione pelo menos um indicador.")

elif pagina == "🏪 Resultado por Loja":
    st.markdown("<div class='section-title'>🏪 Resultado por Loja</div>", unsafe_allow_html=True)
    if resumo_loja.empty:
        st.info("Sem dados por loja.")
    else:
        loja_df = resumo_loja[resumo_loja["Mês"].isin(meses_sel)].copy()
        resumo = loja_df.groupby("Loja", dropna=False)[["Receita", "Despesas", "Resultado Caixa Simplificado"]].sum().reset_index()
        resumo = resumo.sort_values("Resultado Caixa Simplificado", ascending=False)
        fig = px.bar(resumo, x="Loja", y="Resultado Caixa Simplificado", template="plotly_dark", text_auto=".2s")
        fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", height=420)
        st.plotly_chart(fig, use_container_width=True)
        show = resumo.copy()
        for col in ["Receita", "Despesas", "Resultado Caixa Simplificado"]:
            show[col] = show[col].apply(brl)
        st.dataframe(show, use_container_width=True, hide_index=True)

elif pagina == "⚖️ Resultados Estratégicos":
    st.markdown("<div class='section-title'>⚖️ Resultados Estratégicos</div>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        make_kpi("Margem de Contribuição", brl(valor_linha(dados, "MARGEM DE CONTRIBUIÇÃO TOTAL", ultimo_mes)))
    with c2:
        make_kpi("Margem Contribuição %", pct(valor_linha(dados, "MARGEM DE CONTRIBUIÇÃO PERCENTUAL", ultimo_mes)))
    with c3:
        make_kpi("Ponto de Equilíbrio", brl(valor_linha(dados, "PONTO DE EQUILÍBRIO", ultimo_mes)))
    with c4:
        folga = valor_linha(dados, "RECEITA TOTAL", ultimo_mes) - valor_linha(dados, "PONTO DE EQUILÍBRIO", ultimo_mes)
        make_kpi("Folga Operacional", brl(folga))
    sec = dados[dados["Seção"].astype(str).str.upper().str.contains("PONTO", na=False)]
    st.markdown(dre_pivot_html(sec, meses_sel), unsafe_allow_html=True)

elif pagina == "⚠️ Auditoria DRE":
    st.markdown("<div class='section-title'>⚠️ Auditoria DRE</div>", unsafe_allow_html=True)
    if checks is not None and not checks.empty:
        st.dataframe(checks, use_container_width=True, hide_index=True)
    if nao_classificados.empty:
        st.markdown("<div class='ok-box'>Nenhum item não classificado encontrado.</div>", unsafe_allow_html=True)
    else:
        aud = nao_classificados.copy()
        if "Mês" in aud.columns:
            aud = aud[aud["Mês"].isin(meses_sel) | aud["Mês"].astype(str).str.lower().isin(["sem mês", "sem mes"])]
        valor_nc = aud["Valor"].sum() if "Valor" in aud.columns else 0
        qtde_nc = aud["Qtde"].sum() if "Qtde" in aud.columns else len(aud)
        c1, c2 = st.columns(2)
        with c1:
            make_kpi("Valor não classificado", brl(valor_nc))
        with c2:
            make_kpi("Qtde não classificada", f"{int(qtde_nc):,}".replace(",", "."))
        show = aud.copy()
        if "Valor" in show.columns:
            show["Valor"] = show["Valor"].apply(brl)
        st.dataframe(show, use_container_width=True, hide_index=True)

st.markdown("<div class='footer'>EIROX FINANCIAL ANALYTICS • DRE Online Premium • Reprocessamento dinâmico pelas pastas</div>", unsafe_allow_html=True)
